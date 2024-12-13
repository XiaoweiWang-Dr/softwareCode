import csv
import numpy as np
from numpy import unique
from numpy import where
from matplotlib import pyplot
import joblib
from sklearn.neighbors import KNeighborsRegressor
from sklearn.svm import SVR
from sklearn.linear_model import Lasso
from sklearn.ensemble import RandomForestRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.model_selection import train_test_split
import os
grandfa = os.path.dirname(os.path.realpath(__file__))

csv_reader = csv.reader(open(r"C:\data\TempData_A-1_2024-06-13 00_00_00至2024-06-22 23_59_59.csv",encoding='utf-8'))
t = []
t1 = []
t2 = []
t3 = []
for row in csv_reader:
    try:
        t.append(float(row[5]))
        t1.append(float(row[6]))
        t2.append(float(row[7]))
        t3.append(float(row[8]))
        # print(row[5])
    except:
        pass

t = t1+t2+t3

print(len(t))
x_list = []
y_list = []
temp_list = []
y_l_list = []
for i in range(len(t)):
    temp_list.append(t[i])
    if i%10 == 0:
        if i==0:
            temp_list = []
            continue
        x_list.append(np.array(temp_list))
        y_list.append(t[i+1])
        y_temp_list = []
        for j in range(1,91):
            try:
                y_temp_list.append(t[i+j])
            except:
                break
        y_l_list.append(np.array(y_temp_list))
        # print(np.array(y_temp_list))
        temp_list = []
        

for i in range(10):
    y_l_list.pop()
    x_list.pop()
    y_list.pop()
y_l_list = np.array(y_l_list)
data_array = np.concatenate([array[np.newaxis, :] for array in x_list])
y_list = np.array(y_list).reshape(-1,1)
# print(x_list)
# print(y_list)
x_list = data_array
print(x_list.shape)
print(y_list.shape)
print(y_l_list.shape)

X_train, X_test, Y_train, Y_test = train_test_split(x_list,y_list,test_size=0.2,random_state=44)
_X_train, _X_test, Y_l_train, Y_l_test = train_test_split(x_list,y_l_list,test_size=0.2,random_state=44)
print(X_train)
print(_X_train)

clf =RandomForestRegressor()
rf = clf.fit(X_train,Y_train)
y_pred = rf.predict(X_test)

pred = []
labe = []
for i in range(X_test.shape[0]):
    temp_test = X_test[i]
    for j in range(1,90):
        y_pred = rf.predict(temp_test.reshape(1,-1))
        temp_test = np.append(temp_test,y_pred)
        temp_test = np.delete(temp_test,0)
    pred.append(y_pred[0])
    labe.append(Y_l_test[i][89])
    # for j in range(0,10):

# print(X_test.shape)
# print(y_pred.shape)
print("train score：",rf.score(X_train,Y_train))
print("train score：",rf.score(X_test,Y_test))
import matplotlib.pyplot as plt
plt.figure(figsize=(20, 10))
plt.plot(range(int(len(labe)*0.4),int(len(labe)*0.45)), (labe[int(len(labe)*0.4):int(len(labe)*0.45)]), marker='.',label="true value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="predict value")
plt.legend()
# plt.show()
plt.savefig('out90.png',dpi=300)
import openpyxl as p
workbook = p.Workbook()
sheet = workbook.active
sheet.cell(1, 1).value = 'label'
sheet.cell(1, 2).value = 'predict_90'
print(pred)
for i in range(len(labe)):
    sheet.cell(i+2, 1).value = labe[i]
    sheet.cell(i+2, 2).value = pred[i]
workbook.save(r"C:\Users\Alice\Desktop\fs90.xlsx")
print(np.array(labe).shape)
print(np.array(pred).reshape(1,-1)[0].shape)
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
print("mae",mean_absolute_error(np.array(labe),np.array(pred).reshape(1,-1)[0]))
print("mse",mean_squared_error(np.array(labe),np.array(pred).reshape(1,-1)[0]))
print("rmse",np.sqrt(mean_squared_error(np.array(labe),np.array(pred).reshape(1,-1)[0])))