import csv
import numpy as np
from numpy import unique
from numpy import where
from matplotlib import pyplot
import joblib
from sklearn.model_selection import train_test_split
import os
from sklearn.neighbors import KNeighborsRegressor
from sklearn.svm import SVR
from sklearn.linear_model import Lasso
from sklearn.ensemble import RandomForestRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.ensemble import GradientBoostingRegressor

grandfa = os.path.dirname(os.path.realpath(__file__))

csv_reader = csv.reader(open(r"C:\data\TempData_A-1_2024-06-13 00_00_00至2024-06-22 23_59_59.csv",encoding='utf-8'))
t = []
t1 = []
t2 = []
t3 = []
for row in csv_reader:
    try:
        t.append(float(row[5]))
        t1.append(float(row[6]))
        t2.append(float(row[7]))
        t3.append(float(row[8]))
        # print(row[5])
    except:
        pass

t = t1+t2+t3

print(len(t))
x_list = []
y_list = []
temp_list = []
y_l_list = []
for i in range(len(t)):
    temp_list.append(t[i])
    if i%10 == 0:
        if i==0:
            temp_list = []
            continue
        x_list.append(np.array(temp_list))
        y_list.append(t[i+1])
        y_temp_list = []
        for j in range(1,11):
            try:
                y_temp_list.append(t[i+j])
            except:
                break
        y_l_list.append(np.array(y_temp_list))
        # print(np.array(y_temp_list))
        temp_list = []

y_l_list.pop()
x_list.pop()
y_list.pop()
y_l_list = np.array(y_l_list)
data_array = np.concatenate([array[np.newaxis, :] for array in x_list])
y_list = np.array(y_list).reshape(-1,1)
# print(x_list)
# print(y_list)
x_list = data_array
print(x_list.shape)
print(y_list.shape)
print(y_l_list.shape)

X_train, X_test, Y_train, Y_test = train_test_split(x_list,y_list,test_size=0.2,random_state=44)
_X_train, _X_test, Y_l_train, Y_l_test = train_test_split(x_list,y_l_list,test_size=0.2,random_state=44)
print(X_train)
print(_X_train)


knn = KNeighborsRegressor()
svr = SVR()
lasso = Lasso()
clf = RandomForestRegressor()
ab = AdaBoostRegressor()
gb = GradientBoostingRegressor()

knn_ = knn.fit(X_train,Y_train)
svr_ = svr.fit(X_train,Y_train)
lasso_ = lasso.fit(X_train,Y_train)
rf = clf.fit(X_train,Y_train)
ab_ = ab.fit(X_train,Y_train)
gb_ = gb.fit(X_train,Y_train)

# y_pred_knn = knn_.predict(X_test)
# y_pred_svr = svr_.predict(X_test)
# y_pred_lasso = lasso_.predict(X_test)
# y_pred = rf.predict(X_test)
# y_pred_ab = ab_.predict(X_test)
# y_pred_gb = gb_.predict(X_test)

pred_knn = []
labe_knn = []
pred_svr = []
labe_svr = []
pred_lasso = []
labe_lasso = []
pred = []
labe = []
pred_ab = []
labe_ab = []
pred_gb = []
labe_gb = []

for i in range(X_test.shape[0]):

    temp_test = X_test[i]
    for j in range(1, 10):
        y_pred = knn_.predict(temp_test.reshape(1, -1))
        temp_test = np.append(temp_test, y_pred)
        temp_test = np.delete(temp_test, 0)
    pred_knn.append(y_pred[0])
    labe_knn.append(Y_l_test[i][9])

    temp_test = X_test[i]
    for j in range(1, 10):
        y_pred = svr_.predict(temp_test.reshape(1, -1))
        temp_test = np.append(temp_test, y_pred)
        temp_test = np.delete(temp_test, 0)
    pred_svr.append(y_pred[0])
    labe_svr.append(Y_l_test[i][9])


    temp_test = X_test[i]
    for j in range(1, 10):
        y_pred = lasso_.predict(temp_test.reshape(1, -1))
        temp_test = np.append(temp_test, y_pred)
        temp_test = np.delete(temp_test, 0)
    pred_lasso.append(y_pred[0])
    labe_lasso.append(Y_l_test[i][9])

    temp_test = X_test[i]
    for j in range(1,10):
        y_pred = rf.predict(temp_test.reshape(1,-1))
        temp_test = np.append(temp_test,y_pred)
        temp_test = np.delete(temp_test,0)
    pred.append(y_pred[0])
    labe.append(Y_l_test[i][9])

    temp_test = X_test[i]
    for j in range(1, 10):
        y_pred = ab_.predict(temp_test.reshape(1, -1))
        temp_test = np.append(temp_test, y_pred)
        temp_test = np.delete(temp_test, 0)
    pred_ab.append(y_pred[0])
    labe_ab.append(Y_l_test[i][9])

    temp_test = X_test[i]
    for j in range(1, 10):
        y_pred = gb_.predict(temp_test.reshape(1, -1))
        temp_test = np.append(temp_test, y_pred)
        temp_test = np.delete(temp_test, 0)
    pred_gb.append(y_pred[0])
    labe_gb.append(Y_l_test[i][9])
    # for j in range(0,10):

# print(X_test.shape)
# print(y_pred.shape)
print("train score：",rf.score(X_train,Y_train))
print("test score：",rf.score(X_test,Y_test))
import matplotlib.pyplot as plt
plt.figure(figsize=(20, 10))
# plt.figure(dpi=300)
# plt.plot(range(len(labe)), (labe), label="true value")
# plt.plot(range(len(pred)), (pred), label="predict value")
plt.plot(range(int(len(labe)*0.4),int(len(labe)*0.45)), (labe[int(len(labe)*0.4):int(len(labe)*0.45)]), marker='.',label="true value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="RF predict value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred_lasso[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="Lasso predict value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred_knn[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="KNN predict value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred_svr[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="SVR predict value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred_ab[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="AdaBoost predict value")
plt.plot(range(int(len(pred)*0.4),int(len(pred)*0.45)), (pred_gb[int(len(pred)*0.4):int(len(pred)*0.45)]), marker='.',label="GradientBoosting predict value")


plt.legend()
# plt.show()
plt.savefig('out.png',dpi=300)
import openpyxl as p
workbook = p.Workbook()
sheet = workbook.active
sheet.cell(1, 1).value = 'label'
sheet.cell(1, 2).value = 'RF'
sheet.cell(1, 3).value = 'Lasso'
sheet.cell(1, 4).value = 'KNN'
sheet.cell(1, 5).value = 'SVR'
sheet.cell(1, 6).value = 'AdaBoost'
sheet.cell(1, 7).value = 'GradientBoosting'
for i in range(len(labe)):
    sheet.cell(i+2, 1).value = labe[i]
    sheet.cell(i+2, 2).value = pred[i]
    sheet.cell(i + 2, 3).value = pred_lasso[i]
    sheet.cell(i + 2, 4).value = pred_knn[i][0]
    sheet.cell(i + 2, 5).value = pred_svr[i]
    sheet.cell(i + 2, 6).value = pred_ab[i]
    sheet.cell(i + 2, 7).value = pred_gb[i]
workbook.save(r"C:\Users\Alice\Desktop\compare.xlsx")
print(np.array(labe).shape)
print(np.array(pred).reshape(1,-1)[0].shape)
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
print("mae",mean_absolute_error(np.array(labe),np.array(pred).reshape(1,-1)[0]))
print("mse",mean_squared_error(np.array(labe),np.array(pred).reshape(1,-1)[0]))
print("rmse",np.sqrt(mean_squared_error(np.array(labe),np.array(pred).reshape(1,-1)[0])))